"""
Exportador Excel — v4
Gera UMA única planilha com:
  - Aba "DRE": movimentação líquida mensal, mês atual + anterior + AV% + Var.
  - Aba "Balanço Patrimonial": saldo anterior e atual + AV% + Var.
  - Aba "VALIDAÇÃO": erros matemáticos.
  - Aba "EXTRAÇÃO": dump bruto das contas.

Regras:
  - DRE: nunca mostra débito/crédito bruto — apenas movimento líquido calculado.
  - BP: nunca mostra débito/crédito — apenas saldo anterior e saldo atual.
  - AV% DRE: base = Receita Operacional Líquida (após deduções).
  - AV% BP ativo: base = Total Ativo. AV% BP passivo/PL: base = Total Passivo+PL.
  - Resultado do Período apurado na DRE e transferido como linha no PL do BP.
"""
from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import NaturezaSaldo, RelatorioFinal, SeveridadeErro, TipoConta

# ---------------------------------------------------------------------------
# Paleta
# ---------------------------------------------------------------------------
COR_TITULO_BG   = "1F4E78"
COR_TITULO_FG   = "FFFFFF"
COR_GRUPO_BG    = "D9E1F2"
COR_SUBTOTAL_BG = "BDD7EE"
COR_LINHA_PAR   = "F2F2F2"
COR_RESULTADO_BG = "E2EFDA"
COR_PREJUIZO_BG  = "FCE4D6"
COR_ERRO_CRITICO = "FFD7D7"
COR_ERRO_AVISO   = "FFF3CD"

FMT_MOEDA   = '#,##0.00;(#,##0.00);"-"'
FMT_PERCENT = '0%'
FMT_PERCENT2 = '0.00%'


def _slug(texto: str) -> str:
    """Converte nome para slug seguro para nome de arquivo."""
    import unicodedata
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^\w\s-]", "", texto).strip()
    return re.sub(r"[\s]+", "_", texto)[:50]


def _limpar_empresa(nome: str) -> str:
    """
    Remove artefatos de extração do nome da empresa.
    Casos tratados:
    - "Empresa: FRIGORIFICO TAMOYO LTDA Folha: 00"  → "FRIGORIFICO TAMOYO LTDA"
    - "Empresa FRIGORIFICO TAMOYO LTDA Folha: 0001" → "FRIGORIFICO TAMOYO LTDA"
    - "FT TRANSPORTE LTDA"                           → "FT TRANSPORTE LTDA" (inalterado)
    """
    nome = re.sub(r'(?i)^Empresa[:\s]+', '', nome).strip()
    nome = re.sub(r'(?i)\s+Folha[:\s]*\S*\s*$', '', nome).strip()
    return nome


# ---------------------------------------------------------------------------
# Ponto de entrada — gera UMA única planilha
# ---------------------------------------------------------------------------

def exportar(
    relatorio: RelatorioFinal,
    output_dir: str = "output",
    relatorio_anterior: Optional[RelatorioFinal] = None,
) -> str:
    Path(output_dir).mkdir(exist_ok=True)
    empresa   = relatorio.nome_cliente or _limpar_empresa(relatorio.balancete.empresa) or "Empresa"
    slug      = _slug(empresa)
    periodo   = relatorio.balancete.periodo_fim.replace("/", "-") \
                or datetime.now().strftime("%m-%Y")
    sufixo_tp = f"_{relatorio.tipo_pessoa}" if relatorio.tipo_pessoa else ""
    out_path  = f"{output_dir}/Empresa_{slug}{sufixo_tp}_{periodo}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    resultado_atual    = _apurar_resultado(relatorio)
    resultado_anterior = _apurar_resultado(relatorio_anterior) if relatorio_anterior else None

    # DRE primeiro, depois BP (o resultado apurado na DRE alimenta o PL do BP)
    _construir_dre(wb, relatorio, relatorio_anterior, resultado_atual, resultado_anterior)
    _construir_bp(wb, relatorio, relatorio_anterior, resultado_atual)
    _construir_validacao(wb, relatorio)
    _construir_extracao(wb, relatorio)

    wb.save(out_path)
    return out_path


def exportar_sem_modelo(relatorio, output_dir="output", relatorio_anterior=None):
    return exportar(relatorio, output_dir, relatorio_anterior)


# ---------------------------------------------------------------------------
# Movimento líquido e apuração
# ---------------------------------------------------------------------------

def _mov(conta) -> Decimal:
    """Movimento líquido do período para qualquer conta."""
    if conta.natureza == NaturezaSaldo.CREDOR:
        return conta.credito - conta.debito   # receita: mais crédito = positivo
    return conta.debito - conta.credito       # despesa: mais débito = positivo


# ---------------------------------------------------------------------------
# Acumuladores e fórmulas DRE
# ---------------------------------------------------------------------------

_TIPO_SINAL: dict = {
    "receita":             1,
    "deducao":            -1,
    "custo":              -1,
    "despesa":            -1,
    "tributaria":         -1,   # despesas tributárias — separadas para cálculo correto do EBITDA
    "outra_receita":       1,
    "outra_despesa":      -1,
    "depreciacao":         1,
    "receita_financeira":  1,
    "despesa_financeira": -1,
    "provisao_ir":        -1,
}


def _sinal_display(tipo: str) -> int:
    return _TIPO_SINAL.get(tipo, -1)


def _avaliar_formula(formula: str, acc: dict) -> Decimal:
    """Avalia 'receita - deducao - custo' usando acc."""
    resultado = Decimal("0")
    for m in re.finditer(r'([+-]?)\s*([a-z_]+)', formula.strip()):
        sinal = -1 if m.group(1) == '-' else 1
        resultado += sinal * acc.get(m.group(2), Decimal("0"))
    return resultado


def _construir_acc(rel) -> dict:
    """Constrói dicionário {tipo: mov_liquido} a partir de dre_estrutura."""
    from collections import defaultdict
    acc: dict = defaultdict(Decimal)
    if rel and rel.dre_estrutura:
        _acc_estrutura(rel.dre_estrutura, rel, acc)
    return acc


def _acc_estrutura(estrutura: list, rel, acc: dict) -> None:
    for grupo in estrutura:
        tipo  = grupo.get("tipo", "despesa")
        chave = grupo.get("chave", "")
        if tipo == "subtotal":
            val = _avaliar_formula(grupo.get("formula", ""), acc)
            if chave:
                acc[chave] = val
        elif tipo == "agrupador":
            for sub in grupo.get("subgrupos", {}).values():
                mov = _somar_mov_grupo(_pool_dre(rel),
                                       sub.get("prefixos", []))
                # sinal_acc: -1 para contas CREDOR com movimento devedor (ex: CMV em 3.x)
                sinal = Decimal(str(sub.get("sinal_acc", 1)))
                acc[sub.get("tipo", "despesa")] += sinal * mov
        else:
            mov = _somar_mov_grupo(_pool_dre(rel),
                                   grupo.get("prefixos", []),
                                   inverte_credor=grupo.get("inverte_credor", False))
            sinal = Decimal(str(grupo.get("sinal_acc", 1)))
            acc[tipo] += sinal * mov


def _apurar_resultado(rel) -> Decimal:
    """Apura Lucro Liquido somando todos os tipos relevantes do acc."""
    if not rel or not rel.dre_estrutura:
        return _apurar_fallback(rel)
    acc = _construir_acc(rel)
    return (acc.get("receita",              Decimal("0"))
            - acc.get("deducao",            Decimal("0"))
            - acc.get("custo",              Decimal("0"))
            - acc.get("despesa",            Decimal("0"))
            - acc.get("tributaria",         Decimal("0"))
            + acc.get("outra_receita",      Decimal("0"))
            - acc.get("outra_despesa",      Decimal("0"))
            + acc.get("receita_financeira", Decimal("0"))
            - acc.get("despesa_financeira", Decimal("0"))
            - acc.get("provisao_ir",        Decimal("0")))


def _apurar_fallback(rel) -> Decimal:
    resultado = Decimal("0")
    for c in rel.balancete.contas_analiticas:
        if not (c.codigo.startswith("3") or c.codigo.startswith("4")):
            continue
        m = _mov(c)
        resultado += m if c.natureza == NaturezaSaldo.CREDOR else -m
    return resultado


def _somar_mov_grupo(contas, prefixos: list, excluir: list | None = None, inverte_credor: bool = False) -> Decimal:
    excluir = excluir or []
    total = Decimal("0")
    for c in contas:
        if excluir and any(c.codigo.startswith(e) for e in excluir):
            continue
        for p in prefixos:
            if c.codigo.startswith(p):
                mov = _mov(c)
                if inverte_credor and c.natureza == NaturezaSaldo.CREDOR:
                    mov = -mov   # retificadoras CREDOR dentro de grupo DEVEDOR reduzem o total
                total += mov
                break
    return total


def _pool_dre(rel) -> list:
    # ContaCerta usa IDs sequenciais: pai [2884] e filhos [2898],[5694] não
    # compartilham prefixo, então startswith age como match exato — sem double-counting.
    # Contas folha classificadas como SINTETICA (ex: [2037] "Receitas Diversas")
    # ficam fora de contas_analiticas; pesquisar em contas resolve o problema.
    if rel and rel.balancete.layout_detectado == "contacerta":
        return rel.balancete.contas
    return rel.balancete.contas_analiticas


def _contas_do_grupo(todas, prefixos, excluir=None):
    excluir = excluir or []
    out = []
    for c in todas:
        if excluir and any(c.codigo.startswith(e) for e in excluir):
            continue
        for p in prefixos:
            if c.codigo.startswith(p):
                out.append(c)
                break
    return out


def _idx(rel) -> dict:
    """Indice O(1) das analiticas por codigo."""
    return {c.codigo: c for c in rel.balancete.contas_analiticas}


# ---------------------------------------------------------------------------
# ABA: DRE
# ---------------------------------------------------------------------------

def _construir_dre(wb, rel, rel_ant, res_atual, res_anterior):
    ws  = wb.create_sheet("DRE")
    b   = rel.balancete
    com = rel_ant is not None
    idx_ant = _idx(rel_ant) if com else {}
    per_ant = rel_ant.balancete.periodo_fim if com else ""

    # Cabecalho
    n_cols = 7 if com else 4
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    _cel(ws["A1"],
         f"DEMONSTRACAO DO RESULTADO DO EXERCICIO (DRE)  -  "
         f"Empresa: {b.empresa}  Folha: 0001  |  CNPJ: {b.cnpj}",
         bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center", size=11)
    ws.row_dimensions[1].height = 28

    if com:
        ws.merge_cells("C2:D2")
        _cel(ws["C2"], f"{b.periodo_fim}",
             bold=True, bg="2E4057", fg=COR_TITULO_FG, halign="center")
        ws.merge_cells("E2:F2")
        _cel(ws["E2"], f"{per_ant}",
             bold=True, bg="2E4057", fg=COR_TITULO_FG, halign="center")
        for col in ["A", "B", "G"]:
            ws[f"{col}2"].fill = PatternFill("solid", fgColor="2E4057")

    hdrs = [("A3", "DEMONSTRACAO DE RESULTADO"), ("C3", "R$"), ("D3", "AV%")]
    if com:
        hdrs += [("E3", "R$"), ("F3", "AV%"), ("G3", "Var.%")]
    ws.merge_cells("A3:B3")
    for ref, txt in hdrs:
        _cel(ws[ref], txt, bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center")
    ws.row_dimensions[3].height = 18

    row = 4
    rec_liq_atual = _calcular_rec_liq(rel)
    rec_liq_ant   = _calcular_rec_liq(rel_ant) if com else None

    estrutura = rel.dre_estrutura
    if not estrutura:
        row = _dre_fallback(ws, row, rel, idx_ant, com, rec_liq_atual, rec_liq_ant)
    else:
        from collections import defaultdict
        acc     : dict = defaultdict(Decimal)
        acc_ant : dict = defaultdict(Decimal)
        usa_subtotais = any(g.get("tipo") == "subtotal" for g in estrutura)
        ultimo_st = None

        for grupo in estrutura:
            row = _processar_bloco_dre(
                ws, row, grupo, rel, rel_ant if com else None,
                idx_ant, acc, acc_ant, rec_liq_atual, rec_liq_ant, com)

            if not usa_subtotais:
                tp = grupo.get("tipo", "")
                if tp == "deducao" and ultimo_st != "rec_liq":
                    rl_a = float(acc.get("receita", Decimal("0"))
                                 - acc.get("deducao", Decimal("0")))
                    rl_b = float(acc_ant.get("receita", Decimal("0"))
                                 - acc_ant.get("deducao", Decimal("0"))) if com else None
                    row = _linha_dre_subtotal(ws, row, "Receita Operacional Liquida",
                                              rl_a, rl_b, rec_liq_atual, rec_liq_ant, com)
                    ultimo_st = "rec_liq"
                elif tp == "custo" and ultimo_st != "lucro_bruto":
                    lb_a = float(acc.get("receita", Decimal("0"))
                                 - acc.get("deducao", Decimal("0"))
                                 - acc.get("custo",   Decimal("0")))
                    lb_b = float(acc_ant.get("receita", Decimal("0"))
                                 - acc_ant.get("deducao", Decimal("0"))
                                 - acc_ant.get("custo",   Decimal("0"))) if com else None
                    row = _linha_dre_subtotal(ws, row, "Lucro (Prejuizo) Bruto",
                                              lb_a, lb_b, rec_liq_atual, rec_liq_ant, com)
                    ultimo_st = "lucro_bruto"

    row += 1
    label_res = ("Prejuizo Liquido do Periodo"
                 if res_atual < 0 else "Resultado Liquido do Periodo")
    _escrever_linha_resultado(ws, row, label_res,
                              float(res_atual),
                              float(res_anterior) if res_anterior is not None else None,
                              rec_liq_atual, rec_liq_ant, com)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 8
    if com:
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 10
    ws.freeze_panes = "A4"


def _processar_bloco_dre(ws, row, grupo, rel, rel_ant, idx_ant,
                          acc, acc_ant, base_a, base_b, com) -> int:
    tipo  = grupo.get("tipo", "despesa")
    label = grupo.get("label", "")
    chave = grupo.get("chave", "")

    if tipo == "subtotal":
        val_d     = _avaliar_formula(grupo.get("formula", ""), acc)
        val_d_ant = _avaliar_formula(grupo.get("formula", ""), acc_ant) if com else None
        if chave:
            acc[chave] = val_d
            if com and val_d_ant is not None:
                acc_ant[chave] = val_d_ant
        row = _linha_dre_subtotal(ws, row, label,
                                   float(val_d),
                                   float(val_d_ant) if val_d_ant is not None else None,
                                   base_a, base_b, com)

    elif tipo == "agrupador":
        subgrupos = [dict(v, _chave=k) for k, v in grupo.get("subgrupos", {}).items()]
        total_a_d = Decimal("0")
        total_b_d = Decimal("0")
        for sub in subgrupos:
            sub_tipo  = sub.get("tipo", "despesa")
            sub_pref  = sub.get("prefixos", [])
            sub_excl  = sub.get("excluir_prefixos", [])
            sinal_s   = _sinal_display(sub_tipo)
            sinal_acc = Decimal(str(sub.get("sinal_acc", 1)))
            mov = _somar_mov_grupo(_pool_dre(rel), sub_pref, sub_excl)
            total_a_d += sinal_s * mov
            acc[sub_tipo] += sinal_acc * mov   # sinal_acc corrige CREDOR com mov devedor
            if com and rel_ant:
                mov_b = _somar_mov_grupo(_pool_dre(rel_ant), sub_pref, sub_excl)
                total_b_d += sinal_s * mov_b
                acc_ant[sub_tipo] += sinal_acc * mov_b

        row = _linha_dre_grupo(ws, row, label,
                                float(total_a_d),
                                float(total_b_d) if com else None,
                                base_a, base_b, com)

        for sub in subgrupos:
            sub_tipo  = sub.get("tipo", "despesa")
            sub_pref  = sub.get("prefixos", [])
            sub_excl  = sub.get("excluir_prefixos", [])
            sub_label = sub.get("label", "")
            sinal_s   = _sinal_display(sub_tipo)
            sub_mov   = _somar_mov_grupo(_pool_dre(rel), sub_pref, sub_excl)
            sv_a = float(sinal_s * sub_mov)
            sv_b = None
            if com and rel_ant:
                sv_b = float(sinal_s * _somar_mov_grupo(
                    _pool_dre(rel_ant), sub_pref, sub_excl))

            row = _linha_dre_subgrupo(ws, row, sub_label, sv_a, sv_b,
                                       base_a, base_b, com)

            for c in _contas_do_grupo(rel.balancete.contas, sub_pref, sub_excl):
                if c.tipo != TipoConta.ANALITICA:
                    continue
                va = sinal_s * float(_mov(c))
                vb = None
                if com:
                    ca = idx_ant.get(c.codigo)
                    if ca:
                        vb = sinal_s * float(_mov(ca))
                nivel = len(c.codigo.split("."))
                row = _linha_dre_analitica(ws, row, c.codigo, c.descricao,
                                            va, vb, base_a, base_b, com, nivel)

    else:
        prefixos  = grupo.get("prefixos", [])
        inv_cred  = grupo.get("inverte_credor", False)
        sinal     = _sinal_display(tipo)
        mov       = _somar_mov_grupo(_pool_dre(rel), prefixos, inverte_credor=inv_cred)
        val_a     = float(sinal * mov)
        acc[tipo] += mov
        val_b = None
        if com and rel_ant:
            mov_b = _somar_mov_grupo(_pool_dre(rel_ant), prefixos, inverte_credor=inv_cred)
            val_b = float(sinal * mov_b)
            acc_ant[tipo] += mov_b

        row = _linha_dre_grupo(ws, row, label, val_a, val_b, base_a, base_b, com)

        for c in _contas_do_grupo(rel.balancete.contas, prefixos):
            if c.tipo != TipoConta.ANALITICA:
                continue
            va = sinal * float(_mov(c))
            vb = None
            if com:
                ca = idx_ant.get(c.codigo)
                if ca:
                    vb = sinal * float(_mov(ca))
            nivel = len(c.codigo.split("."))
            row = _linha_dre_analitica(ws, row, c.codigo, c.descricao,
                                        va, vb, base_a, base_b, com, nivel)

    return row


def _calcular_rec_liq(rel) -> float:
    """Receita Operacional Liquida = Receita Bruta - Deducoes (base AV%).
    Retorna 0.0 quando nao ha receita — _av interpreta base==0 como None (exibe '-').
    """
    if rel is None or not rel.dre_estrutura:
        return 0.0
    acc = _construir_acc(rel)
    val = float(acc.get("receita", Decimal("0")) - acc.get("deducao", Decimal("0")))
    return val  # 0.0 quando zero → _av retorna None → celula exibe "-"


def _av(valor, base):
    if valor is None or base == 0:
        return None
    return valor / abs(base)


def _var_r(atual, anterior):
    if atual is None or anterior is None:
        return None
    return atual - anterior


def _var_p(atual, anterior):
    if atual is None or anterior is None or anterior == 0:
        return None
    return (atual - anterior) / abs(anterior)


def _linha_dre_grupo(ws, row, label, val_a, val_b, base_a, base_b, com) -> int:
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=COR_GRUPO_BG)
    c.alignment = Alignment(vertical="center")
    _val_cel(ws, row, 3, val_a, bold=True, bg=COR_GRUPO_BG)
    _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT, bold=True, bg=COR_GRUPO_BG)
    if com:
        _val_cel(ws, row, 5, val_b, bold=True, bg=COR_GRUPO_BG)
        _val_cel(ws, row, 6, _av(val_b, base_b), fmt=FMT_PERCENT, bold=True, bg=COR_GRUPO_BG)
        _val_cel(ws, row, 7, _var_p(val_a, val_b), fmt=FMT_PERCENT, bold=True, bg=COR_GRUPO_BG)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COR_GRUPO_BG)
    return row + 1


def _linha_dre_subgrupo(ws, row, label, val_a, val_b, base_a, base_b, com) -> int:
    COR_SUB = "E3ECFA"
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=COR_SUB)
    c.alignment = Alignment(vertical="center", indent=1)
    _val_cel(ws, row, 3, val_a, bold=True, bg=COR_SUB)
    _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT, bold=True, bg=COR_SUB)
    if com:
        _val_cel(ws, row, 5, val_b, bold=True, bg=COR_SUB)
        _val_cel(ws, row, 6, _av(val_b, base_b), fmt=FMT_PERCENT, bold=True, bg=COR_SUB)
        _val_cel(ws, row, 7, _var_p(val_a, val_b), fmt=FMT_PERCENT, bold=True, bg=COR_SUB)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COR_SUB)
    return row + 1


def _linha_dre_analitica(ws, row, cod, desc, val_a, val_b,
                          base_a, base_b, com, nivel) -> int:
    bg = COR_LINHA_PAR if row % 2 == 0 else "FFFFFF"
    ws.cell(row=row, column=1, value=cod).font = Font(name="Arial", size=9)
    cd = ws.cell(row=row, column=2, value=desc)
    cd.font      = Font(name="Arial", size=9)
    cd.alignment = Alignment(indent=max(1, nivel - 2))
    _val_cel(ws, row, 3, val_a, bg=bg if bg != "FFFFFF" else None, size=9)
    _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT,
             bg=bg if bg != "FFFFFF" else None, size=9)
    if com:
        _val_cel(ws, row, 5, val_b, bg=bg if bg != "FFFFFF" else None, size=9)
        _val_cel(ws, row, 6, _av(val_b, base_b), fmt=FMT_PERCENT,
                 bg=bg if bg != "FFFFFF" else None, size=9)
        _val_cel(ws, row, 7, _var_p(val_a, val_b), fmt=FMT_PERCENT,
                 bg=bg if bg != "FFFFFF" else None, size=9)
    if bg != "FFFFFF":
        for col in [1, 2]:
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
    return row + 1


def _linha_dre_subtotal(ws, row, label, val_a, val_b, base_a, base_b, com) -> int:
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, name="Arial", size=10)
    c.fill      = PatternFill("solid", fgColor=COR_SUBTOTAL_BG)
    c.alignment = Alignment(vertical="center")
    _val_cel(ws, row, 3, val_a, bold=True, bg=COR_SUBTOTAL_BG)
    _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT, bold=True, bg=COR_SUBTOTAL_BG)
    if com:
        _val_cel(ws, row, 5, val_b, bold=True, bg=COR_SUBTOTAL_BG)
        _val_cel(ws, row, 6, _av(val_b, base_b), fmt=FMT_PERCENT, bold=True, bg=COR_SUBTOTAL_BG)
        _val_cel(ws, row, 7, _var_p(val_a, val_b), fmt=FMT_PERCENT, bold=True, bg=COR_SUBTOTAL_BG)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=COR_SUBTOTAL_BG)
    return row + 1


def _escrever_linha_resultado(ws, row, label, val_a, val_b, base_a, base_b, com):
    bg = COR_RESULTADO_BG if val_a >= 0 else COR_PREJUIZO_BG
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(bold=True, name="Arial", size=10, color=COR_TITULO_BG)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center")
    _val_cel(ws, row, 3, val_a, bold=True, bg=bg)
    _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT, bold=True, bg=bg)
    if com:
        _val_cel(ws, row, 5, val_b, bold=True, bg=bg)
        _val_cel(ws, row, 6, _av(val_b, base_b), fmt=FMT_PERCENT, bold=True, bg=bg)
        _val_cel(ws, row, 7, _var_p(val_a, val_b), fmt=FMT_PERCENT, bold=True, bg=bg)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)


def _dre_fallback(ws, row, rel, idx_ant, com, base_a, base_b):
    contas = [c for c in rel.balancete.contas
              if c.codigo.startswith("3") or c.codigo.startswith("4")]
    for c in contas:
        nivel = len(c.codigo.split("."))
        mv = float(_mov(c))
        va = mv if c.natureza == NaturezaSaldo.CREDOR else -mv
        vb = None
        if com:
            ca = idx_ant.get(c.codigo)
            if ca:
                mvb = float(_mov(ca))
                vb = mvb if ca.natureza == NaturezaSaldo.CREDOR else -mvb
        if c.tipo == TipoConta.SINTETICA or nivel <= 2:
            row = _linha_dre_grupo(ws, row, c.descricao, va, vb,
                                    base_a, base_b, com)
        else:
            row = _linha_dre_analitica(ws, row, c.codigo, c.descricao,
                                        va, vb, base_a, base_b, com, nivel)
    return row


# ---------------------------------------------------------------------------
# ABA: Balanço Patrimonial — YAML-driven (ContaCerta e layouts sem hierarquia 1.x/2.x)
# ---------------------------------------------------------------------------

# Ordem canônica das seções do BP e seus títulos de exibição
_SECOES_BP_YAML = [
    ("ativo_circulante",       "ATIVO CIRCULANTE"),
    ("ativo_nao_circulante",   "ATIVO NÃO CIRCULANTE"),
    ("passivo_circulante",     "PASSIVO CIRCULANTE"),
    ("passivo_nao_circulante", "PASSIVO NÃO CIRCULANTE"),
    ("patrimonio_liquido",     "PATRIMÔNIO LÍQUIDO"),
]


def _display_grupo_bp(grupo, secao: str) -> float:
    """
    Converte o valor interno do mapper para valor de exibição no BP.

    Convenção do mapper: DEVEDOR → +saldo, CREDOR → -saldo.
    - Ativo: +D = ativo normal, -C = saldo credor (cheque especial etc.) → usar como está.
    - Passivo/PL normais (sinal=1): CREDOR → mapper gerou -saldo → inverter para exibição positiva.
    - Passivo/PL redutores (sinal=-1, ex: Prejuízo): valor já negativo por design → manter negativo.
    """
    val = float(grupo.valor)
    if secao.startswith("ativo"):
        return val
    # passivo / patrimônio_liquido
    if grupo.sinal == -1:
        return val   # conta redutora do PL: já é negativo, manter
    return -val      # inversão: CREDOR foi negado pelo mapper, desfazer para exibição


def _construir_bp_yaml(wb, rel: RelatorioFinal,
                       rel_ant: Optional[RelatorioFinal],
                       resultado_periodo: Decimal):
    """
    Renderiza o Balanço Patrimonial a partir de rel.grupos_balanco (mapeamento YAML).
    Usado para layouts cujos códigos não seguem a convenção hierárquica 1.x/2.x (ex: ContaCerta).
    """
    ws  = wb.create_sheet("Balanço Patrimonial")
    b   = rel.balancete
    com = rel_ant is not None
    resultado_anterior = _apurar_resultado(rel_ant) if com else None
    eh_prej     = resultado_periodo < Decimal("0")
    eh_prej_ant = (resultado_anterior < Decimal("0")) if resultado_anterior is not None else False

    # Totais para base do AV%
    total_ativo_a = sum(
        abs(_display_grupo_bp(g, k.split(".")[0]))
        for k, g in rel.grupos_balanco.items()
        if k.split(".")[0].startswith("ativo")
    ) or 1.0
    total_passpl_a = sum(
        abs(_display_grupo_bp(g, k.split(".")[0]))
        for k, g in rel.grupos_balanco.items()
        if not k.split(".")[0].startswith("ativo")
    ) or 1.0
    total_ativo_a  += abs(float(resultado_periodo)) if not eh_prej else 0
    total_passpl_a += abs(float(resultado_periodo))

    if com and rel_ant.grupos_balanco:
        total_ativo_b = sum(
            abs(_display_grupo_bp(g, k.split(".")[0]))
            for k, g in rel_ant.grupos_balanco.items()
            if k.split(".")[0].startswith("ativo")
        ) or 1.0
        total_passpl_b = sum(
            abs(_display_grupo_bp(g, k.split(".")[0]))
            for k, g in rel_ant.grupos_balanco.items()
            if not k.split(".")[0].startswith("ativo")
        ) or 1.0
        if resultado_anterior is not None:
            total_passpl_b += abs(float(resultado_anterior))
    else:
        total_ativo_b = total_passpl_b = 1.0

    # Título
    n_cols = 8 if com else 5
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    _cel(ws["A1"],
         f"BALANÇO PATRIMONIAL  —  {b.empresa}   |   CNPJ: {b.cnpj}",
         bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center", size=11)
    ws.row_dimensions[1].height = 28

    # Cabeçalho colunas
    if com:
        hdrs = ["Conta", "Descrição",
                f"Saldo {rel_ant.balancete.periodo_fim}", "AV%",
                f"Saldo {b.periodo_fim}", "AV%",
                "Var. R$", "Var. %"]
    else:
        hdrs = ["Conta", "Descrição", f"Saldo {b.periodo_fim}", "AV%", "Var. %"]

    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color=COR_TITULO_FG, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=COR_TITULO_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    row = 3

    for secao_key, secao_label in _SECOES_BP_YAML:
        grupos_secao = {
            k: g for k, g in rel.grupos_balanco.items()
            if k.startswith(secao_key + ".")
        }
        if not grupos_secao:
            continue

        base_a = total_ativo_a if secao_key.startswith("ativo") else total_passpl_a
        base_b = (total_ativo_b if secao_key.startswith("ativo") else total_passpl_b) if com else None

        # Cabeçalho da seção
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row=row, column=1, value=secao_label)
        c.font      = Font(bold=True, name="Arial", size=10, color=COR_TITULO_FG)
        c.fill      = PatternFill("solid", fgColor=COR_TITULO_BG)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

        total_secao_a: float = 0.0
        total_secao_b = None

        # Resultado do Período — inserido no início do PL
        if secao_key == "patrimonio_liquido":
            bg_res     = COR_PREJUIZO_BG if eh_prej     else COR_RESULTADO_BG
            bg_res_ant = COR_PREJUIZO_BG if eh_prej_ant else COR_RESULTADO_BG
            label_res  = "(-) Resultado do Período (Prejuízo)" if eh_prej else "Resultado do Período"
            val_res_a  = float(resultado_periodo)
            val_res_b  = float(resultado_anterior) if resultado_anterior is not None else None

            ws.cell(row=row, column=1, value="RESULT.")
            cd = ws.cell(row=row, column=2, value=label_res)
            cd.alignment = Alignment(indent=1)
            if com:
                _val_cel(ws, row, 3, val_res_b, bold=True, bg=bg_res_ant)
                _val_cel(ws, row, 4, _av(val_res_b, base_b), fmt=FMT_PERCENT, bold=True, bg=bg_res_ant)
                _val_cel(ws, row, 5, val_res_a, bold=True, bg=bg_res)
                _val_cel(ws, row, 6, _av(val_res_a, base_a), fmt=FMT_PERCENT, bold=True, bg=bg_res)
                _val_cel(ws, row, 7, _var_r(val_res_a, val_res_b), bold=True, bg=bg_res)
                _val_cel(ws, row, 8, _var_p(val_res_a, val_res_b), fmt=FMT_PERCENT, bold=True, bg=bg_res)
            else:
                _val_cel(ws, row, 3, val_res_a, bold=True, bg=bg_res)
                _val_cel(ws, row, 4, _av(val_res_a, base_a), fmt=FMT_PERCENT, bold=True, bg=bg_res)
            for col in [1, 2]:
                c = ws.cell(row=row, column=col)
                c.font      = Font(bold=True, name="Arial", size=10, color=COR_TITULO_BG)
                c.fill      = PatternFill("solid", fgColor=bg_res)
                c.alignment = Alignment(vertical="center")
            row += 1

            total_secao_a += val_res_a
            if com and val_res_b is not None:
                total_secao_b = (total_secao_b or 0.0) + val_res_b

        # Grupos da seção
        for chave, grupo in grupos_secao.items():
            val_a = _display_grupo_bp(grupo, secao_key)
            val_b = None
            if com and rel_ant.grupos_balanco:
                grupo_ant = rel_ant.grupos_balanco.get(chave)
                if grupo_ant:
                    val_b = _display_grupo_bp(grupo_ant, secao_key)

            bg = COR_LINHA_PAR if row % 2 == 0 else "FFFFFF"
            ws.cell(row=row, column=1, value="")
            cd = ws.cell(row=row, column=2, value=grupo.label or grupo.nome)
            cd.font      = Font(name="Arial", size=10)
            cd.alignment = Alignment(indent=1)

            if com:
                _val_cel(ws, row, 3, val_b, bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 4, _av(val_b, base_b), fmt=FMT_PERCENT,
                         bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 5, val_a, bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 6, _av(val_a, base_a), fmt=FMT_PERCENT,
                         bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 7, _var_r(val_a, val_b),
                         bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 8, _var_p(val_a, val_b), fmt=FMT_PERCENT,
                         bg=bg if bg != "FFFFFF" else None)
            else:
                _val_cel(ws, row, 3, val_a, bg=bg if bg != "FFFFFF" else None)
                _val_cel(ws, row, 4, _av(val_a, base_a), fmt=FMT_PERCENT,
                         bg=bg if bg != "FFFFFF" else None)

            if bg != "FFFFFF":
                for col in [1, 2]:
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
            for col in [1, 2]:
                ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

            total_secao_a += val_a
            if com and val_b is not None:
                total_secao_b = (total_secao_b or 0.0) + val_b
            row += 1

        # Subtotal da seção
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=f"TOTAL {secao_label}")
        c.font      = Font(bold=True, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=COR_SUBTOTAL_BG)
        c.alignment = Alignment(vertical="center")
        if com:
            _val_cel(ws, row, 3, total_secao_b, bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 4, _av(total_secao_b, base_b), fmt=FMT_PERCENT,
                     bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 5, total_secao_a, bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 6, _av(total_secao_a, base_a), fmt=FMT_PERCENT,
                     bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 7, _var_r(total_secao_a, total_secao_b),
                     bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 8, _var_p(total_secao_a, total_secao_b), fmt=FMT_PERCENT,
                     bold=True, bg=COR_SUBTOTAL_BG)
        else:
            _val_cel(ws, row, 3, total_secao_a, bold=True, bg=COR_SUBTOTAL_BG)
            _val_cel(ws, row, 4, _av(total_secao_a, base_a), fmt=FMT_PERCENT,
                     bold=True, bg=COR_SUBTOTAL_BG)
        row += 1

    # Dimensões
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    if com:
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 10
    else:
        ws.column_dimensions["E"].width = 10
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# ABA: Balanço Patrimonial — legado (Socium e layouts com hierarquia 1.x/2.x)
# ---------------------------------------------------------------------------

def _construir_bp(wb, rel: RelatorioFinal,
                  rel_ant: Optional[RelatorioFinal],
                  resultado_periodo: Decimal):
    # Layouts sem hierarquia numérica 1.x/2.x usam o BP guiado pelo YAML (grupos_balanco).
    # Layouts hierárquicos (Socium) continuam usando o renderer legado por prefixo.
    if rel.grupos_balanco and rel.balancete.layout_detectado == "contacerta":
        _construir_bp_yaml(wb, rel, rel_ant, resultado_periodo)
        return

    ws  = wb.create_sheet("Balanço Patrimonial")
    b   = rel.balancete
    com = rel_ant is not None

    # Resultado do mês anterior — usado para incorporar no PL da coluna anterior
    resultado_anterior = _apurar_resultado(rel_ant) if com else None

    # Calcular totais para AV%
    total_ativo_atual   = _total_grupo(rel, ["1"])
    total_passivo_atual = _total_grupo(rel, ["2"])
    total_ativo_ant     = _total_grupo(rel_ant, ["1"]) if com else None
    total_passivo_ant   = _total_grupo(rel_ant, ["2"]) if com else None

    # ── Título ────────────────────────────────────────────────────────────
    n_cols = 8 if com else 5
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    _cel(ws["A1"],
         f"BALANÇO PATRIMONIAL  —  {b.empresa}   |   CNPJ: {b.cnpj}",
         bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center", size=11)
    ws.row_dimensions[1].height = 28

    # ── Cabeçalho colunas ─────────────────────────────────────────────────
    if com:
        hdrs = ["Conta", "Descrição",
                f"Saldo {rel_ant.balancete.periodo_fim}", "AV%",
                f"Saldo {b.periodo_fim}", "AV%",
                "Var. R$", "Var. %"]
    else:
        hdrs = ["Conta", "Descrição", f"Saldo {b.periodo_fim}", "AV%", "Var. %"]

    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color=COR_TITULO_FG, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=COR_TITULO_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Corpo ─────────────────────────────────────────────────────────────
    contas_bp = [c for c in b.contas
                 if c.codigo.startswith("1") or c.codigo.startswith("2")]

    row = 3
    resultado_inserido     = False
    resultado_ant_inserido = False
    eh_prej     = resultado_periodo   < Decimal("0")
    eh_prej_ant = (resultado_anterior < Decimal("0")) if resultado_anterior is not None else False

    for conta in contas_bp:
        nivel   = len(conta.codigo.split("."))
        eh_sint = conta.tipo == TipoConta.SINTETICA

        # Base AV%: ativo → total_ativo; passivo/PL → total_passivo
        base_a = total_ativo_atual if conta.codigo.startswith("1") else total_passivo_atual
        base_b = (total_ativo_ant  if conta.codigo.startswith("1") else total_passivo_ant) if com else None

        sal_atual  = _saldo_bp(conta.saldo_atual, conta)

        # ── Saldo da coluna anterior (Opção B) ────────────────────────────
        # Para contas do Ativo e Passivo Circulante: saldo_atual do balancete anterior
        # Para contas do PL: saldo_atual do balancete anterior
        # O resultado do mês anterior é inserido como linha própria no PL
        sal_ant_bp = None
        if com:
            conta_ant = next(
                (c for c in rel_ant.balancete.contas if c.codigo == conta.codigo), None
            )
            sal_ant_bp = _saldo_bp(conta_ant.saldo_atual, conta_ant) if conta_ant else None

        # ── Inserir linha Resultado do Período ATUAL ──────────────────────
        # Posição: imediatamente antes do primeiro grupo Lucros/Prejuízos do PL
        if not resultado_inserido and _e_grupo_lucros(conta):
            row = _inserir_resultado_bp(
                ws, row,
                val_atual=float(resultado_periodo),
                val_ant=float(resultado_anterior) if resultado_anterior is not None else None,
                eh_prej_atu=eh_prej,
                eh_prej_ant=eh_prej_ant,
                com=com,
                base_a=base_a,
                base_b=base_b,
            )
            resultado_inserido     = True
            resultado_ant_inserido = True

        row = _linha_bp(ws, row, conta, sal_ant_bp, sal_atual,
                        base_a, base_b, eh_sint, nivel, com)

    if not resultado_inserido:
        _inserir_resultado_bp(
            ws, row,
            val_atual=float(resultado_periodo),
            val_ant=float(resultado_anterior) if resultado_anterior is not None else None,
            eh_prej_atu=eh_prej,
            eh_prej_ant=eh_prej_ant,
            com=com,
            base_a=total_passivo_atual,
            base_b=total_passivo_ant,
        )

    # ── Dimensões ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    if com:
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 10
    else:
        ws.column_dimensions["E"].width = 10
    ws.freeze_panes = "A3"


def _total_grupo(rel: Optional[RelatorioFinal], prefixos: list[str]) -> float:
    if rel is None:
        return 1.0
    total = sum(
        abs(float(c.saldo_atual))
        for c in rel.balancete.contas_analiticas
        if any(c.codigo.startswith(p) for p in prefixos)
    )
    return total if total != 0 else 1.0


def _e_grupo_lucros(conta) -> bool:
    """
    Detecta o ponto de inserção do Resultado no PL.
    Dispara na primeira conta SINTÉTICA de nível 2 ou 3 dentro do PL
    cujo código seja 2.3.x ou 2.4.x E cuja descrição indique
    Lucros/Prejuízos Acumulados.
    Não dispara no cabeçalho geral do PL (2.3 ou 2.4 puro).
    """
    cod = conta.codigo
    # Deve ser subgrupo do PL (ao menos 3 segmentos: 2.3.x ou 2.4.x)
    partes = cod.split(".")
    if len(partes) < 3:
        return False
    if not (cod.startswith("2.3") or cod.startswith("2.4")):
        return False
    desc = conta.descricao.lower()
    return any(p in desc for p in
               ["lucro", "prejuizo", "prejuízo", "resultado", "acumulado"])


def _inserir_resultado_bp(ws, row, val_atual, val_ant,
                          eh_prej_atu, eh_prej_ant,
                          com, base_a, base_b) -> int:
    """
    Insere a linha 'Resultado do Período' dentro do PL do BP.

    - val_atual: resultado apurado do mês atual (positivo=lucro, negativo=prejuízo)
    - val_ant:   resultado apurado do mês anterior (Opção B: já incorporado no PL)
    - Cor da célula: verde (lucro) ou laranja (prejuízo), baseada no mês atual
    - Ambas as colunas preenchidas quando com=True
    """
    label = ("(-) Resultado do Período (Prejuízo)"
             if eh_prej_atu else "Resultado do Período")
    bg = COR_PREJUIZO_BG if eh_prej_atu else COR_RESULTADO_BG

    ws.cell(row=row, column=1, value="RESULT.")
    ws.cell(row=row, column=2, value=label)

    if com:
        # Coluna anterior: resultado do mês anterior (já apurado da DRE anterior)
        bg_ant = COR_PREJUIZO_BG if eh_prej_ant else COR_RESULTADO_BG
        _val_cel(ws, row, 3, val_ant,  bold=True, bg=bg_ant)
        _val_cel(ws, row, 4, _av(val_ant, base_b), fmt=FMT_PERCENT, bold=True, bg=bg_ant)
        # Coluna atual
        _val_cel(ws, row, 5, val_atual, bold=True, bg=bg)
        _val_cel(ws, row, 6, _av(val_atual, base_a), fmt=FMT_PERCENT, bold=True, bg=bg)
        # Variações
        _val_cel(ws, row, 7, _var_r(val_atual, val_ant), bold=True, bg=bg)
        _val_cel(ws, row, 8, _var_p(val_atual, val_ant), fmt=FMT_PERCENT, bold=True, bg=bg)
    else:
        _val_cel(ws, row, 3, val_atual, bold=True, bg=bg)
        _val_cel(ws, row, 4, _av(val_atual, base_a), fmt=FMT_PERCENT, bold=True, bg=bg)

    for col in [1, 2]:
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, name="Arial", size=10, color=COR_TITULO_BG)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center")
    return row + 1


def _linha_bp(ws, row, conta, sal_ant, sal_atu,
              base_a, base_b, eh_sint, nivel, com) -> int:
    bg   = COR_GRUPO_BG if (eh_sint or nivel <= 2) else (COR_LINHA_PAR if row % 2 == 0 else "FFFFFF")
    bold = eh_sint or nivel <= 2

    ws.cell(row=row, column=1, value=conta.codigo)
    cd = ws.cell(row=row, column=2, value=conta.descricao)
    cd.alignment = Alignment(indent=max(0, nivel - 1))

    if com:
        _val_cel(ws, row, 3, sal_ant, bold=bold, bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 4, _av(sal_ant, base_b), fmt=FMT_PERCENT, bold=bold,
                 bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 5, sal_atu, bold=bold, bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 6, _av(sal_atu, base_a), fmt=FMT_PERCENT, bold=bold,
                 bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 7, _var_r(sal_atu, sal_ant), bold=bold,
                 bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 8, _var_p(sal_atu, sal_ant), fmt=FMT_PERCENT, bold=bold,
                 bg=bg if bg != "FFFFFF" else None)
    else:
        _val_cel(ws, row, 3, sal_atu, bold=bold, bg=bg if bg != "FFFFFF" else None)
        _val_cel(ws, row, 4, _av(sal_atu, base_a), fmt=FMT_PERCENT, bold=bold,
                 bg=bg if bg != "FFFFFF" else None)

    for col in [1, 2]:
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=bold, name="Arial", size=10)
        if bg != "FFFFFF":
            c.fill = PatternFill("solid", fgColor=bg)

    return row + 1


def _saldo_bp(saldo: Decimal, conta) -> float:
    val = float(abs(saldo))
    # Ativo (1.x): natureza normal = DEVEDOR; saldo credor = invertido → negativo (parênteses)
    if conta.codigo.startswith("1"):
        return -val if conta.natureza == NaturezaSaldo.CREDOR else val
    # Passivo/PL (2.x): natureza normal = CREDOR; saldo devedor = invertido → negativo (parênteses)
    if conta.codigo.startswith("2"):
        return -val if conta.natureza == NaturezaSaldo.DEVEDOR else val
    return val


# ---------------------------------------------------------------------------
# ABA: Validação
# ---------------------------------------------------------------------------

def _construir_validacao(wb, rel: RelatorioFinal):
    ws = wb.create_sheet("VALIDAÇÃO")
    ws.merge_cells("A1:G1")
    _cel(ws["A1"], "RELATÓRIO DE VALIDAÇÃO CONTÁBIL",
         bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center")

    for col, h in enumerate(
        ["Severidade","Conta","Descrição","Mensagem",
         "Valor Esperado","Valor Encontrado","Diferença"], 1
    ):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=COR_TITULO_FG, name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
        c.alignment = Alignment(horizontal="center")

    erros = rel.balancete.erros_validacao
    if not erros:
        ws.cell(row=3, column=1,
                value="✓ Nenhum erro contábil encontrado.").font = \
            Font(bold=True, color="2D6A4F", name="Arial")
        return

    for i, e in enumerate(erros, 3):
        cor = COR_ERRO_CRITICO if e.severidade == SeveridadeErro.CRITICO else COR_ERRO_AVISO
        for col, val in enumerate([
            e.severidade.value, e.conta_codigo, e.conta_descricao, e.mensagem,
            float(e.valor_esperado)   if e.valor_esperado   else "",
            float(e.valor_encontrado) if e.valor_encontrado else "",
            float(e.diferenca)        if e.diferenca        else "",
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=cor)
            c.font = Font(name="Arial", size=10)

    for col, w in zip("ABCDEFG", [12,15,40,70,18,18,15]):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# ABA: Extração bruta
# ---------------------------------------------------------------------------

def _construir_extracao(wb, rel: RelatorioFinal):
    ws = wb.create_sheet("EXTRAÇÃO")
    b  = rel.balancete

    ws.merge_cells("A1:H1")
    _cel(ws["A1"],
         f"EXTRAÇÃO BRUTA — {b.empresa}  |  CNPJ: {b.cnpj}  |  "
         f"{b.periodo_inicio} a {b.periodo_fim}",
         bold=True, bg=COR_TITULO_BG, fg=COR_TITULO_FG, halign="center")

    meta = [
        ("Sistema Contábil", b.sistema_contabil),
        ("Layout Detectado", b.layout_detectado),
        ("Total de Contas",  len(b.contas)),
        ("Contas Analíticas",len(b.contas_analiticas)),
        ("Contas Sintéticas",len(b.contas_sinteticas)),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=i, column=2, value=val).font  = Font(name="Arial", size=10)

    hr = len(meta) + 3
    for col, h in enumerate(
        ["Código","Descrição","Tipo","Natureza",
         "Sal. Anterior","Débito","Crédito","Sal. Atual"], 1
    ):
        c = ws.cell(row=hr, column=col, value=h)
        c.font = Font(bold=True, color=COR_TITULO_FG, name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
        c.alignment = Alignment(horizontal="center")

    for i, conta in enumerate(b.contas, hr + 1):
        for col, val in enumerate([
            conta.codigo, conta.descricao, conta.tipo.value, conta.natureza.value,
            float(conta.saldo_anterior), float(conta.debito),
            float(conta.credito), float(conta.saldo_atual),
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=COR_LINHA_PAR)
        for col in range(5, 9):
            ws.cell(row=i, column=col).number_format = FMT_MOEDA

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    for col in ["E","F","G","H"]:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = f"A{hr+1}"


# ---------------------------------------------------------------------------
# Utilitários gerais
# ---------------------------------------------------------------------------

def _val_cel(ws, row, col, valor, bold=False, bg=None, fg="000000",
             fmt=FMT_MOEDA, size=10):
    if valor is None:
        c = ws.cell(row=row, column=col, value="-")
        c.font = Font(name="Arial", size=size, color="AAAAAA")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center")
        return
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, name="Arial", size=size, color=fg)
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)


def _cel(cel, valor, bold=False, bg=None, fg="000000",
         halign=None, size=10):
    cel.value     = valor
    cel.font      = Font(bold=bold, color=fg, name="Arial", size=size)
    cel.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    if bg:
        cel.fill = PatternFill("solid", fgColor=bg)
    return cel
